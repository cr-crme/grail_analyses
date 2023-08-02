# -*- coding: utf-8 -*-
"""
Created on Wed Jul 26 11:32:50 2023

@author: Florence
"""

import os
import sys
import datetime
import glob
import numpy as np
import inspect

from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QPushButton, QLineEdit, QGridLayout, QMessageBox

import jeux
import calculs
from formatting import set_border, set_alignment, optimize_cell_width, bold, no_bold


def Excel(ID, file, directory):
    games = [name for name, obj in inspect.getmembers(
        jeux) if inspect.isfunction(obj)]
    # Creation d'un feuille "Presentation" si le fichier Excel n'existait pas deja
    if os.path.exists(file) == False:
        wb = Workbook()
        ws = wb.create_sheet("Presentation")
        del wb['Sheet']

        ws["A1"] = "ID"
        ws["B1"] = ID
        ws["A2"] = "Nom"
        ws["A3"] = "Prenom"
        ws["A4"] = "Numéro de dossier"
        ws["A5"] = "Date de création"
        ws["A6"] = "Derniere date de modification"
        ws["B5"] = datetime.datetime.today().strftime('%Y-%m-%d')

        for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
            for cell in row:
                bold(ws, cell, 'v2')
                pass

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 15
        wb.save(file)

    wb = load_workbook(file)
    for game in games:
        fonction = getattr(jeux, game)
        info = fonction()
        motif_recherche = f'{directory}/{ID}/Record Data/Fichiers txt/{info[0]}'
        fichiers_correspondants = glob.glob(motif_recherche, recursive=True)

        if len(fichiers_correspondants) != 0:  # Creation d'une nouvelle feuille si necessaire
            if info[1] in wb.sheetnames:
                ws = wb[info[1]]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.font and cell.font.bold:
                            no_bold(ws, cell, 'v2')
            else:
                ws = wb.create_sheet(info[1])

            # Ecriture du nom des variables dans Excel
            for l in range(len(info[2])):
                ws[chr(65+l)+str(1)] = info[2][l]
                bold(ws, chr(65+l)+str(1))

            i = 2
            for fichier in fichiers_correspondants:
                data = np.genfromtxt(fichier, delimiter="\t", skip_header=1)

                nomFichier = os.path.basename(fichier)
                partiesNom = nomFichier.split('_')
                AAAA = partiesNom[2][0:4]
                MM = partiesNom[2][4:6]
                DD = partiesNom[2][6:8]
                calculResultats = getattr(calculs, game)
                resultats = calculResultats(info[1], data, AAAA, MM, DD)

                for l in range(len(resultats)):
                    ws[chr(65+l)+str(i)] = resultats[l]
                    ws[chr(65+l)+str(i)].number_format = "0.00"
                i += 1

            ws[chr(65)+str(i+1)] = "Signature :"
            bold(ws, chr(65)+str(i+1))

            # Mise en page
            cells_to_border_1, cells_to_border_2, cells_to_border_3, cells_to_border_4, cells_to_border_5 = [], [], [], [], []

            for col in range(len(info[2])):
                cells_to_border_1.append(chr(65+col) + str(1))
                cells_to_border_4.append(chr(65+col) + str(1))
                cells_to_border_4.append(chr(65+col) + str(i))

            for row in range(1, i):
                cells_to_border_2.append("A" + str(row))
                cells_to_border_5.append(chr(64+len(info[2])) + str(row))
                for col in range(len(info[2])):
                    if row > 1 and row < i:
                        cells_to_border_3.append(chr(65+col) + str(row))

            set_border(ws, cells_to_border_1, "top", "thick")
            set_border(ws, cells_to_border_2, "left", "thick")
            set_border(ws, cells_to_border_3, "bottom", "thin")
            set_border(ws, cells_to_border_4, "bottom", "thick")
            set_border(ws, cells_to_border_5, "right", "thick")

            set_alignment(ws, "center")
            set_alignment(ws, "left", ["A"+str(i+1)])
            optimize_cell_width(ws)
    wb.save(file)


def check_ID():
    # MODIFIER SELON LE REPERTOIRE UTILISE POUR LES ENREGISTREMENTS DES FICHIERS. Probablement D:\CAREN Ressources\Data
    directory = "C:/Users/Florence/Desktop/GRAIL SIMULATION"
    ID = ID_edit.text()
    file = directory + "/" + ID + \
        f"/Record Data/{ID}_ResumeIntervention.xlsx"  # Fichier cree

    # Verifier si l'ID est valide
    for root, subfolders, files in os.walk(directory):
        if ID in subfolders:
            break
        else:
            QMessageBox.warning(window, 'Erreur', 'Enter a valid ID.')
            ID_edit.clear()
            return

    if os.path.exists(file):  # Verifier si le document est modifiable
        try:
            with open(file, 'a+'):
                pass
            Excel(ID, file, directory)
        except IOError:
            QMessageBox.warning(window, 'Validation Error',
                                'File is currently open. Please close the file.')
            ID_edit.clear()
            return
    else:
        Excel(ID, file, directory)
    window.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = QWidget()
    window.setWindowTitle('Recherche de patients')
    window.setGeometry(500, 500, 300, 20)

    layout = QGridLayout()

    ID_label = QLabel('ID patient', window)
    layout.addWidget(ID_label, 0, 0)

    ID_edit = QLineEdit(window)
    layout.addWidget(ID_edit, 0, 1)

    search_button = QPushButton('Terminer', window)
    search_button.clicked.connect(check_ID)
    layout.addWidget(search_button, 4, 0, 1, 2)
    ID_edit.returnPressed.connect(search_button.click)

    window.setLayout(layout)
    window.show()

    sys.exit(app.exec_())
