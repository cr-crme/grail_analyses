# -*- coding: utf-8 -*-
"""
Created on Tue Aug  1 13:58:40 2023

@author: Florence
"""

from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


def set_border(ws, cells, side, thickness):
    for cell in cells:
        existing_border = ws[cell].border
        left_border = existing_border.left
        right_border = existing_border.right
        top_border = existing_border.top
        bottom_border = existing_border.bottom
        new_border = Border(top=top_border if side != 'bottom' else Side(border_style=thickness, color="000000"),
                            bottom=bottom_border if side != 'top' else Side(
                                border_style=thickness, color="000000"),
                            left=left_border if side != 'left' else Side(
                                border_style=thickness, color="000000"),
                            right=right_border if side != 'right' else Side(border_style=thickness, color="000000"))
        ws[cell].border = new_border


def set_alignment(ws, alignment_type, cell_range=None):
    alignment = Alignment(horizontal=alignment_type, vertical="center")
    if cell_range:
        for cell in cell_range:
            ws[cell].alignment = alignment
    else:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment


def optimize_cell_width(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 5
        ws.column_dimensions[get_column_letter(
            col[0].column)].width = adjusted_width


def bold(ws, cell, version='v1'):
    if version == 'v1':
        ws[cell].font = Font(bold=True)
    else:
        cell.font = Font(bold=True)


def no_bold(ws, cell, version='v1'):
    if version == 'v1':
        ws[cell].font = Font(bold=False)
    else:
        cell.font = Font(bold=False)
