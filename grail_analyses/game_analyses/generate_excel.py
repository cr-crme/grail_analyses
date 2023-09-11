import datetime
import glob
import os

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

from .games import get_games_list


def generate_excel(patient_id, save_path):
    # Add an introduction sheet if there were none
    if not os.path.exists(save_path):
        _initialize_file(patient_id, save_path)
    wb = load_workbook(save_path)

    save_folder = os.path.dirname(save_path)
    # Create each game sheet
    for game in get_games_list():
        files = glob.glob(f"{save_folder}/{game.save_name}", recursive=True)
        if not files:
            continue

        if game.name in wb.sheetnames:
            wb.remove(wb[game.name])
        ws = wb.create_sheet(game.name)

        results = None
        i = -1
        for i, file in enumerate(files, start=2):
            data = np.genfromtxt(file, delimiter="\t", skip_header=1)
            aaaa, mm, dd = _get_date(file_path=file)

            results = game.results(data=data, date=f"{aaaa}-{mm}-{dd}")
            for col in range(len(results)):
                ws[f"{chr(65 + col)}{i}"] = results[col][1]
                ws[f"{chr(65 + col)}{i}"].number_format = "0.00"

        ws[f"{chr(65)}{len(files) + 3}"] = "Signature :"
        _bold(ws, f"{chr(65)}{len(files) + 3}")

        for col in range(len(results)):  # Header
            ws[f"{chr(65 + col)}1"] = results[col][0]
            _bold(ws, f"{chr(65 + col)}1")

        # Mise en page
        border1, border2, border3, border4, border5 = [], [], [], [], []

        for col in range(len(results)):
            border1.append(f"{chr(65 + col)}1")
            border4.append(f"{chr(65 + col)}1")
            border4.append(f"{chr(65 + col)}{i}")

        for row in range(1, i):
            border2.append(f"A{row}")
            border5.append(f"{chr(64 + len(results))}{row}")
            for col in range(len(results)):
                if 1 < row < i:
                    border3.append(f"{chr(65 + col)}{row}")

        _set_border(ws, border1, "top", "thick")
        _set_border(ws, border2, "left", "thick")
        _set_border(ws, border3, "bottom", "thin")
        _set_border(ws, border4, "bottom", "thick")
        _set_border(ws, border5, "right", "thick")

        _set_alignment(ws, "center")
        _set_alignment(ws, "left", [f"A{i + 1}"])

        _optimize_cell_width(ws)
    wb.save(save_path)


def _get_date(file_path: str):
    date = os.path.basename(file_path).split('_')
    aaaa = date[2][0:4]
    mm = date[2][4:6]
    dd = date[2][6:8]
    return aaaa, mm, dd


def _initialize_file(patient_id: str, file_path):
    wb = Workbook()
    ws = wb.create_sheet("Presentation")
    del wb['Sheet']

    ws["A1"] = "ID"
    ws["B1"] = patient_id
    ws["A2"] = "Nom"
    ws["A3"] = "Prénom"
    ws["A4"] = "Numéro de dossier"
    ws["A5"] = "Date de création"
    ws["A6"] = "Dernière date de modification"
    ws["B5"] = datetime.datetime.today().strftime('%Y-%m-%d')

    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
        for cell in row:
            _bold(ws, cell, 'v2')

    _optimize_cell_width(ws)
    wb.save(file_path)


def _set_border(ws, cells, side, thickness):
    for cell in cells:
        existing_border = ws[cell].border
        left_border = existing_border.left
        right_border = existing_border.right
        top_border = existing_border.top
        bottom_border = existing_border.bottom
        new_border = Border(top=top_border if side != 'bottom' else Side(border_style=thickness, color="000000"),
                            bottom=bottom_border if side != 'top' else Side(
                                border_style=thickness, color="000000"),
                            left=left_border if side != 'left' else Side(
                                border_style=thickness, color="000000"),
                            right=right_border if side != 'right' else Side(border_style=thickness, color="000000"))
        ws[cell].border = new_border


def _set_alignment(ws, alignment_type, cell_range=None):
    alignment = Alignment(horizontal=alignment_type, vertical="center")
    if cell_range:
        for cell in cell_range:
            ws[cell].alignment = alignment
    else:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment


def _optimize_cell_width(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value is not None and len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width


def _bold(ws, cell, version='v1'):
    if version == 'v1':
        ws[cell].font = Font(bold=True)
    else:
        cell.font = Font(bold=True)


def _not_bold(ws, cell, version='v1'):
    if version == 'v1':
        ws[cell].font = Font(bold=False)
    else:
        cell.font = Font(bold=False)
