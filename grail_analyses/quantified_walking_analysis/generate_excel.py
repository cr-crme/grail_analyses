from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .config import movement_names


def _populate_table(sheet, data, name: str, level: str, plane: str, initial_row: int, initial_col: str) -> int:
    def col(_i):
        return chr(ord(initial_col) + _i)

    if level not in ("kinematics", "moment", "power"):
        raise ValueError("Wrong level")
    if plane not in ("sagittal", "frontal", "transversal"):
        raise ValueError("Wrong plane")

    names = movement_names[level][plane]

    sheet[f"{col(0)}{initial_row}"] = name
    for i, value in enumerate(names):
        sheet[f"{col(0)}{i + initial_row + 1}"] = value["fr"]
        sheet[f"{col(1)}{i + initial_row + 1}"] = round(data[names[i]["en"]]["Minimum Appui"], 2)
        sheet[f"{col(2)}{i + initial_row + 1}"] = round(data[names[i]["en"]]["Maximum Appui"], 2)
        sheet[f"{col(3)}{i + initial_row + 1}"] = round(data[names[i]["en"]]["Range Appui"], 2)
        sheet[f"{col(4)}{i + initial_row + 1}"] = round(data[names[i]["en"]]["Minimum Oscillation"], 2)
        sheet[f"{col(5)}{i + initial_row + 1}"] = round(data[names[i]["en"]]["Maximum Oscillation"], 2)
        sheet[f"{col(6)}{i + initial_row + 1}"] = round(data[names[i]["en"]]["Range Oscillation"], 2)

    # Formatting
    sheet[f"{col(0)}{initial_row}"].font = Font(bold=True)

    # Return the last row
    return initial_row + len(movement_names[level][plane])


def generate_excel(data: dict, save_file: str):
    workbook = Workbook()
    sheet = workbook.active

    _prepare_main_header(sheet)

    last_row = {}
    for initial_col, level in zip(("B", "J", "R"), ("kinematics", "moment", "power")):
        initial_row = 8
        last_row[level] = {}
        for name, plane in zip(("Sagittal", "Frontal", "Transversal"), ("sagittal", "frontal", "transversal")):
            last_row[level][plane] = _populate_table(
                sheet,
                data=data,
                name=name,
                level=level,
                plane=plane,
                initial_row=initial_row,
                initial_col=initial_col,
            )
            initial_row = last_row[level][plane] + 1

    # Center text in the cell
    alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    for column in sheet.columns:  # Adjust cell dimension
        sheet.column_dimensions[column[0].column_letter].width = 12
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["J"].width = 40
        sheet.column_dimensions["R"].width = 40

    _add_borders(sheet, last_row)

    workbook.save(save_file)


def _add_borders(sheet, last_row: dict):
    def to_col(col_name, i):
        return chr(ord(col_name) + i)

    thick = Side(border_style="thick", color="000000")
    med = Side(border_style="medium", color="000000")
    for col, level in zip(("B", "J", "R"), ("kinematics", "moment", "power")):
        row = last_row[level]
        _square_borders(
            sheet, thick, top_left=f"{to_col(col, 0)}5", bottom_right=f"{to_col(col, 6)}{row['transversal']}"
        )
        _square_borders(sheet, thick, top_left=f"{to_col(col, 1)}5", bottom_right=f"{to_col(col, 6)}7")
        _square_borders(
            sheet,
            med,
            top_left=f"{to_col(col, 0)}{row['sagittal'] + 1}",
            bottom_right=f"{to_col(col, 6)}{row['frontal']}",
        )
        _square_borders(
            sheet, thick, top_left=f"{to_col(col, 1)}5", bottom_right=f"{to_col(col, 6)}{row['transversal']}"
        )
        _square_borders(
            sheet, thick, top_left=f"{to_col(col, 0)}8", bottom_right=f"{to_col(col, 3)}{row['transversal']}"
        )


def _square_borders(sheet, format, top_left: str, bottom_right: str):
    top_right = f"{bottom_right[0]}{top_left[1:]}"
    bottom_left = f"{top_left[0]}{bottom_right[1:]}"

    for s in sheet[f"{top_left}:{top_right}"][0]:
        s.border = Border(top=format, left=s.border.left, right=s.border.right, bottom=s.border.bottom)
    for s in sheet[f"{bottom_left}:{bottom_right}"][0]:
        s.border = Border(top=s.border.top, left=s.border.left, right=s.border.right, bottom=format)
    for s in sheet[f"{top_left}:{bottom_left}"]:
        s[0].border = Border(top=s[0].border.top, left=format, right=s[0].border.right, bottom=s[0].border.bottom)
    for s in sheet[f"{top_right}:{bottom_right}"]:
        s[0].border = Border(top=s[0].border.top, left=s[0].border.left, right=format, bottom=s[0].border.bottom)


def _prepare_main_header(sheet):
    sheet["B2"] = "Patient ID"
    sheet["B3"] = "Session ID"
    sheet["B4"] = "Date"
    sheet["C5"] = "Cinématique"
    sheet["C6"] = "Phase d'appui"
    sheet["F6"] = "Phase d'oscillation"
    sheet["C7"] = "Min (deg)"
    sheet["D7"] = "Max (deg)"
    sheet["E7"] = "Range(deg)"
    sheet["F7"] = "Min (deg)"
    sheet["G7"] = "Max (deg)"
    sheet["H7"] = "Range(deg)"

    sheet["K5"] = "Moment"
    sheet["K6"] = "Phase d'appui"
    sheet["N6"] = "Phase d'oscillation"
    sheet["K7"] = "Min (deg)"
    sheet["L7"] = "Max (deg)"
    sheet["M7"] = "Range(deg)"
    sheet["N7"] = "Min (deg)"
    sheet["O7"] = "Max (deg)"
    sheet["P7"] = "Range(deg)"

    sheet["S5"] = "Power"
    sheet["S6"] = "Phase d'appui"
    sheet["V6"] = "Phase d'oscillation"
    sheet["S7"] = "Min (deg)"
    sheet["T7"] = "Max (deg)"
    sheet["U7"] = "Range(deg)"
    sheet["V7"] = "Min (deg)"
    sheet["W7"] = "Max (deg)"
    sheet["X7"] = "Range(deg)"

    # Formatting
    format = Side(border_style="thick", color="000000")
    _square_borders(sheet, format=format, top_left="B2", bottom_right="H4")

    sheet.merge_cells("C2:H2")
    sheet.merge_cells("C3:H3")
    sheet.merge_cells("C4:H4")

    sheet.merge_cells("C5:H5")
    sheet["C5"].font = Font(bold=True)
    sheet.merge_cells("C6:E6")
    sheet.merge_cells("F6:H6")
    sheet["C6"].font = Font(bold=True)
    sheet["F6"].font = Font(bold=True)

    sheet.merge_cells("K5:P5")
    sheet["K5"].font = Font(bold=True)
    sheet.merge_cells("K6:M6")
    sheet.merge_cells("N6:P6")
    sheet["K6"].font = Font(bold=True)
    sheet["N6"].font = Font(bold=True)

    sheet.merge_cells("S5:X5")
    sheet["S5"].font = Font(bold=True)
    sheet.merge_cells("S6:U6")
    sheet.merge_cells("V6:X6")
    sheet["S6"].font = Font(bold=True)
    sheet["V6"].font = Font(bold=True)
