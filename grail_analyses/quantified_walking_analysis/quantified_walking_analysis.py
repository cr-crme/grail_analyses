# -*- coding: utf-8 -*-
"""
Created on Thu Jul 6 14:57:52 2023

@author: Florence
"""
import os
import sys
import io
import math

import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QTextEdit, QVBoxLayout, QHBoxLayout, \
    QWidget, QLabel

# PDF
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak
from reportlab.lib import colors
from PyPDF2 import PdfFileWriter, PdfFileReader

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from config import movement_names

global Files
Files = []


def ranges_of_motion(file_data, variable, side, length):
    if side not in ("left", "right"):
        raise ValueError("Wrong side")

    data = file_data[file_data.iloc[:, 1] == variable]
    data_side = data[data.iloc[:, 0] == side]
    average = data_side.iloc[:, 3:].mean()

    swing_phase = average[length[0 if side == "left" else 1]:]
    maximum_swing = max(swing_phase)
    minimum_swing = min(swing_phase)
    range_swing = maximum_swing - minimum_swing

    stance_phase = average[:length[0 if side == "left" else 1]]
    maximum_stance = max(stance_phase)
    minimum_stance = min(stance_phase)
    range_stance = maximum_stance - minimum_stance

    return {
        "Maximum Oscillation": maximum_swing,
        "Minimum Oscillation": minimum_swing,
        "Range Oscillation": range_swing,
        "Maximum Appui": maximum_stance,
        "Minimum Appui": minimum_stance,
        "Range Appui": range_stance,
    }


def create_page_numbered_pdf(file_path):
    existing_pdf = PdfFileReader(open(file_path, "rb"))
    num_pages = existing_pdf.getNumPages()
    output = PdfFileWriter()
    for i in range(num_pages):
        page = existing_pdf.getPage(i)

        # Create a canvas and add page number
        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=letter)

        can.drawString(520, 20, f"{i + 1} of {num_pages}")
        can.drawString(50, 20, "Signature:____________________")

        can.save()

        packet.seek(0)
        new_pdf = PdfFileReader(packet)
        page.mergePage(new_pdf.getPage(0))

        # Add the page to the output pdf
        output.addPage(page)

    # Write the output pdf
    with open(file_path, "wb") as output_stream:
        output.write(output_stream)


def find_stance_len(data, side) -> int:
    if side not in ("left", "right"):
        raise ValueError("Wrong side")

        # Find number of columns for swing and stance phase
    swing_time = data[data.iloc[:, 1] == f"{'L' if side == 'left' else 'R'}.Swing.Time"]
    stance_time = data[data.iloc[:, 1] == f"{'L' if side == 'left' else 'R'}.Stance.Time"]

    sided_swing_time = swing_time[swing_time.iloc[:, 0] == side]
    sided_stance_time = stance_time[stance_time.iloc[:, 0] == side]

    swing_time_average = sided_swing_time.iloc[:, 3].mean()
    stance_time_average = sided_stance_time.iloc[:, 3].mean()

    total_time_average = swing_time_average + stance_time_average
    swing_percent = swing_time_average / total_time_average * 100
    return math.ceil(swing_percent * 101 / 100)


def _populate_excel_data(
    excel_data: dict, data: pd.DataFrame, level: str, plane: str, side: str, n_swings: tuple[int, int]
):
    if level not in ("kinematics", "moment", "power"):
        raise ValueError("Wrong level")
    if plane not in ("sagittal", "frontal", "transversal"):
        raise ValueError("Wrong plane")
    if side not in ("left", "right"):
        raise ValueError("Wrong side")

    names = movement_names[level][plane]
    for index in range(0 if side == "left" else 1, len(names), 2):
        name = names[index]["en"]
        excel_data[name] = ranges_of_motion(data, name, side, n_swings)


def _populate_measures(measures: list, excel_data: dict, name: str, level: str, plane: str):
    if level not in ("kinematics", "moment", "power"):
        raise ValueError("Wrong level")
    if plane not in ("sagittal", "frontal", "transversal"):
        raise ValueError("Wrong plane")

    measures.append([name])
    names = movement_names[level][plane]
    for i in range(len(names)):
        row = [
            names[i]["fr"],
            round(excel_data[names[i]["en"]]["Minimum Appui"], 2),
            round(excel_data[names[i]["en"]]["Maximum Appui"], 2),
            round(excel_data[names[i]["en"]]["Range Appui"], 2),
            round(excel_data[names[i]["en"]]["Minimum Oscillation"], 2),
            round(excel_data[names[i]["en"]]["Maximum Oscillation"], 2),
            round(excel_data[names[i]["en"]]["Range Oscillation"], 2),
        ]
        measures.append(row)


def process_files():
    for file in Files:
        excel_data = {}
        data = pd.read_csv(file, header=None)

        n_swings = find_stance_len(data, "left"), find_stance_len(data, "right")
        # Kinematics
        for level in ("kinematics", "moment", "power"):
            for side in ("left", "right"):
                for plane in ("sagittal", "frontal", "transversal"):
                    _populate_excel_data(excel_data, data=data, level=level, plane=plane, side=side, n_swings=n_swings)

        # PDF
        folder_path = os.path.dirname(file) + "/"  # À MODIFIER
        file_name_pdf = os.path.splitext(os.path.basename(file))[0] + ".pdf"  # NOM DU FICHIER

        doc = SimpleDocTemplate(
            folder_path + file_name_pdf, pagesize=letter, leftMargin=0.5 * inch, rightMargin=0.5 * inch,
            topMargin=0.4 * inch, bottomMargin=0.5 * inch)
        elements = []

        identification = [["Numéro de dossier"], ["Nom", ""], ["Prénom", ""], ["Date", ""]]
        identification_table = Table(identification, colWidths=[106, 400])
        identification_table.setStyle(
            TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
        )
        elements.append(identification_table)
        elements.append(Spacer(1, 20))  # Add an empty spacer for separation

        # Cinématique
        title = [["Cinématique"]]
        title_table = Table(title, colWidths=[506])
        title_table.setStyle(
            TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ])
        )
        elements.append(title_table)

        phases = [["", "Phase d'Appui", "Phase d'Oscillations"]]
        phases_table = Table(phases, colWidths=[140, 183, 183])
        phases_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ]))
        elements.append(phases_table)

        measures = [["", "Min (deg)", "Max (deg)", "Range (deg)", "Min (deg)", "Max (deg)", "Range (deg)"]]

        for name, plane in zip(("Sagittal", "Frontal", "Transversal"), ("sagittal", "frontal", "transversal")):
            _populate_measures(measures, excel_data=excel_data, name=name, level="kinematics", plane=plane)

        n_sagittal = len(movement_names["kinematics"]["sagittal"]) - 1
        n_frontal = len(movement_names["kinematics"]["frontal"]) - 1
        measure_table = Table(measures, colWidths=[140, 61, 61, 61, 61, 61, 61])
        measure_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, n_sagittal + 3), (0, n_sagittal + 3), 'Helvetica-Bold'),
            ('FONTNAME', (0, n_sagittal + n_frontal + 5), (0, n_sagittal + n_frontal + 5), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (-4, 0), (-4, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, n_sagittal + 2), (-1, n_sagittal + 2), 1, colors.black),
            ('LINEBELOW', (0, n_sagittal + n_frontal + 4), (-1, n_sagittal + n_frontal + 4), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
        ]))

        elements.append(measure_table)
        elements.append(PageBreak())

        # Moment
        title = [["Moment"]]
        title_table = Table(title, colWidths=[475])
        title_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(title_table)

        phases = [["", "Phase d'Appui", "Phase d'Oscillations"]]
        phases_table = Table(phases, colWidths=[140, 183, 183])
        phases_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ]))
        elements.append(phases_table)

        measures = [["", "Min (deg)", "Max (deg)", "Range (deg)", "Min (deg)", "Max (deg)", "Range (deg)"]]
        for name, plane in zip(("Sagittal", "Frontal", "Transversal"), ("sagittal", "frontal", "transversal")):
            _populate_measures(measures, excel_data=excel_data, name=name, level="moment", plane=plane)

        measure_table = Table(measures, colWidths=[140, 61, 61, 61, 61, 61, 61])
        n_sagittal = len(movement_names["moment"]["sagittal"]) - 1
        n_frontal = len(movement_names["moment"]["frontal"]) - 1
        measure_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, n_sagittal + 3), (0, n_sagittal + 3), 'Helvetica-Bold'),
            ('FONTNAME', (0, n_sagittal + n_frontal + 5), (0, n_sagittal + n_frontal + 5), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (-4, 0), (-4, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, n_sagittal + 2), (-1, n_sagittal + 2), 1, colors.black),
            ('LINEBELOW', (0, n_sagittal + n_frontal + 4), (-1, n_sagittal + n_frontal + 4), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
        ]))

        elements.append(measure_table)
        elements.append(PageBreak())

        # Power
        title = [["Puissance"]]
        title_table = Table(title, colWidths=[475])
        title_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(title_table)

        phases = [["", "Phase d'Appui", "Phase d'Oscillations"]]
        phases_table = Table(phases, colWidths=[140, 183, 183])
        phases_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
        ]))
        elements.append(phases_table)

        measures = [["", "Min (deg)", "Max (deg)", "Range (deg)", "Min (deg)", "Max (deg)", "Range (deg)"]]

        for name, plane in zip(("Sagittal", "Frontal", "Transversal"), ("sagittal", "frontal", "transversal")):
            _populate_measures(measures, excel_data=excel_data, name=name, level="power", plane=plane)

        measure_table = Table(measures, colWidths=[140, 61, 61, 61, 61, 61, 61])
        n_sagittal = len(movement_names["power"]["sagittal"]) - 1
        n_frontal = len(movement_names["power"]["frontal"]) - 1
        measure_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, n_sagittal + 3), (0, n_sagittal + 3), 'Helvetica-Bold'),
            ('FONTNAME', (0, n_sagittal + n_frontal + 5), (0, n_sagittal + n_frontal + 5), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (-4, 0), (-4, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, n_sagittal + 2), (-1, n_sagittal + 2), 1, colors.black),
            ('LINEBELOW', (0, n_sagittal + n_frontal + 4), (-1, n_sagittal + n_frontal + 4), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
        ]))

        elements.append(measure_table)
        elements.append(PageBreak())

        doc.build(elements)
        create_page_numbered_pdf(folder_path + file_name_pdf)

        # EXCEL
        folder_path = os.path.dirname(file) + "/"
        file_name_xlsx = os.path.splitext(os.path.basename(file))[0] + ".xlsx"

        workbook = Workbook()
        sheet = workbook.active
        # Titre de cellules
        sheet["B2"] = "Patient ID"
        sheet["B3"] = "Session ID"
        sheet["B4"] = "Date"
        sheet["C5"] = "Cinématique"
        sheet["C6"] = "Phase d'appui"
        sheet["F6"] = "Phase d'oscillation"
        sheet["C7"] = "Min (deg)"
        sheet["D7"] = "Max (deg)"
        sheet["E7"] = "Range(deg)"
        sheet["F7"] = "Min (deg)"
        sheet["G7"] = "Max (deg)"
        sheet["H7"] = "Range(deg)"

        sheet["K5"] = "Moment"
        sheet["K6"] = "Phase d'appui"
        sheet["N6"] = "Phase d'oscillation"
        sheet["K7"] = "Min (deg)"
        sheet["L7"] = "Max (deg)"
        sheet["M7"] = "Range(deg)"
        sheet["N7"] = "Min (deg)"
        sheet["O7"] = "Max (deg)"
        sheet["P7"] = "Range(deg)"

        sheet["S5"] = "Power"
        sheet["S6"] = "Phase d'appui"
        sheet["V6"] = "Phase d'oscillation"
        sheet["S7"] = "Min (deg)"
        sheet["T7"] = "Max (deg)"
        sheet["U7"] = "Range(deg)"
        sheet["V7"] = "Min (deg)"
        sheet["W7"] = "Max (deg)"
        sheet["X7"] = "Range(deg)"

        # Titre des variables

        # Kinematics
        i1 = 8
        sheet["B" + str(i1)] = "Sagittal"

        count = 0
        names = movement_names["kinematics"]["sagittal"]
        for i1, value in enumerate(names, start=i1 + 1):
            sheet["B" + str(i1)] = value["fr"]
            sheet["C" + str(i1)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["D" + str(i1)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["E" + str(i1)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["F" + str(i1)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["G" + str(i1)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["H" + str(i1)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        j1 = i1 + 1
        sheet["B" + str(j1)] = "Frontal"

        count = 0
        names = movement_names["kinematics"]["frontal"]
        for j1, value in enumerate(names, start=(j1 + 1)):
            sheet["B" + str(j1)] = value["fr"]
            sheet["C" + str(j1)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["D" + str(j1)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["E" + str(j1)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["F" + str(j1)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["G" + str(j1)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["H" + str(j1)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        k1 = j1 + 1
        sheet["B" + str(k1)] = "Transversal"

        count = 0
        names = movement_names["kinematics"]["transversal"]
        for k1, value in enumerate(names, start=(k1 + 1)):
            sheet["B" + str(k1)] = value["fr"]
            sheet["C" + str(k1)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["D" + str(k1)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["E" + str(k1)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["F" + str(k1)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["G" + str(k1)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["H" + str(k1)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        # Moment
        i2 = 8
        sheet["J" + str(i2)] = "Sagittal"

        count = 0
        names = movement_names["moment"]["sagittal"]
        for i2, value in enumerate(names, start=i2 + 1):
            sheet["J" + str(i2)] = value["fr"]
            sheet["K" + str(i2)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["L" + str(i2)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["M" + str(i2)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["N" + str(i2)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["O" + str(i2)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["P" + str(i2)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        j2 = i2 + 1
        sheet["J" + str(j2)] = "Frontal"

        count = 0
        j2 = i2 + 1
        names = movement_names["moment"]["frontal"]
        for j2, value in enumerate(names, start=(j2 + 1)):
            sheet["J" + str(j2)] = value["fr"]
            sheet["K" + str(j2)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["L" + str(j2)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["M" + str(j2)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["N" + str(j2)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["O" + str(j2)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["P" + str(j2)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        k2 = j2 + 1
        sheet["J" + str(k2)] = "Transversal"

        count = 0
        names = movement_names["moment"]["transversal"]
        for k2, value in enumerate(names, start=(k2 + 1)):
            sheet["J" + str(k2)] = value["fr"]
            sheet["K" + str(k2)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["L" + str(k2)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["M" + str(k2)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["N" + str(k2)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["O" + str(k2)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["P" + str(k2)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        # Power
        i3 = 8
        sheet["R" + str(i3)] = "Sagittal"

        count = 0
        names = movement_names["power"]["sagittal"]
        for i3, value in enumerate(names, start=i3 + 1):
            sheet["R" + str(i3)] = value["fr"]
            sheet["S" + str(i3)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["T" + str(i3)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["U" + str(i3)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["V" + str(i3)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["W" + str(i3)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["X" + str(i3)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        j3 = i3 + 1
        sheet["R" + str(j3)] = "Frontal"

        count = 0
        j3 = i3 + 1
        names = movement_names["power"]["frontal"]
        for j3, value in enumerate(names, start=(j3 + 1)):
            sheet["R" + str(j3)] = value["fr"]
            sheet["S" + str(j3)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["T" + str(j3)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["U" + str(j3)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["V" + str(j3)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["W" + str(j3)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["X" + str(j3)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        k3 = j3 + 1
        sheet["R" + str(k3)] = "Transversal"

        count = 0
        names = movement_names["power"]["transversal"]
        for k3, value in enumerate(names, start=(k3 + 1)):
            sheet["R" + str(k3)] = value["fr"]
            sheet["S" + str(k3)] = round(excel_data[names[count]["en"]]["Minimum Appui"], 2)
            sheet["T" + str(k3)] = round(excel_data[names[count]["en"]]["Maximum Appui"], 2)
            sheet["U" + str(k3)] = round(excel_data[names[count]["en"]]["Range Appui"], 2)
            sheet["V" + str(k3)] = round(excel_data[names[count]["en"]]["Minimum Oscillation"], 2)
            sheet["W" + str(k3)] = round(excel_data[names[count]["en"]]["Maximum Oscillation"], 2)
            sheet["X" + str(k3)] = round(excel_data[names[count]["en"]]["Range Oscillation"], 2)
            count += 1

        # Mise en forme
        sheet.merge_cells("C2:H2")  # Merge de cellules
        sheet.merge_cells("C3:H3")
        sheet.merge_cells("C4:H4")
        sheet.merge_cells("C5:H5")
        sheet.merge_cells("K5:P5")
        sheet.merge_cells("S5:X5")
        sheet.merge_cells("C6:E6")
        sheet.merge_cells("F6:H6")
        sheet.merge_cells("K6:M6")
        sheet.merge_cells("N6:P6")
        sheet.merge_cells("S6:U6")
        sheet.merge_cells("V6:X6")

        sheet["C5"].font = Font(bold=True)  # Mise de titres en gras
        sheet["K5"].font = Font(bold=True)
        sheet["S5"].font = Font(bold=True)
        sheet["C6"].font = Font(bold=True)
        sheet["F6"].font = Font(bold=True)
        sheet["K6"].font = Font(bold=True)
        sheet["N6"].font = Font(bold=True)
        sheet["S6"].font = Font(bold=True)
        sheet["V6"].font = Font(bold=True)
        sheet["B8"].font = Font(bold=True)  # Sagittal
        sheet["J8"].font = Font(bold=True)
        sheet["R8"].font = Font(bold=True)
        sheet["B" + str(i1 + 1)].font = Font(bold=True)  # Frontal
        sheet["J" + str(i2 + 1)].font = Font(bold=True)
        sheet["R" + str(i3 + 1)].font = Font(bold=True)
        sheet["B" + str(j1 + 1)].font = Font(bold=True)  # Transversal
        sheet["J" + str(j2 + 1)].font = Font(bold=True)
        sheet["R" + str(j3 + 1)].font = Font(bold=True)

        # Center text in the cell
        alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = alignment

        cells_to_border_1 = []  # Borders
        cells_to_border_2 = []
        cells_to_border_3 = []
        border_1 = Border(top=Side(border_style="thick", color="000000"))
        border_2 = Border(left=Side(border_style="thick", color="000000"))
        border_3 = Border(bottom=Side(border_style="medium", color="000000"))

        # First table
        for l in [2, 5, 8, k1 + 1]:
            for m in range(2, 9):
                cells_to_border_1.append(chr(64 + m) + str(l))

        for l in range(2, k1 + 1):
            for m in ["B", "I"]:
                cells_to_border_2.append(m + str(l))

        for l in range(5, k1 + 1):
            cells_to_border_2.append("C" + str(l))

        for l in range(8, k1 + 1):
            cells_to_border_2.append("F" + str(l))

        for m in range(2, 9):
            cells_to_border_3.append(chr(64 + m) + str(i1))
            cells_to_border_3.append(chr(64 + m) + str(j1))

        # Second table
        for l in [5, 8, k2 + 1]:
            for m in range(10, 17):
                cells_to_border_1.append(chr(64 + m) + str(l))

        for l in range(5, k2 + 1):
            for m in ["J", "K", "Q"]:
                cells_to_border_2.append(m + str(l))

        for l in range(8, k2 + 1):
            cells_to_border_2.append("N" + str(l))

        for m in range(10, 17):
            cells_to_border_3.append(chr(64 + m) + str(i2))
            cells_to_border_3.append(chr(64 + m) + str(j2))

        # Third table
        for l in [5, 8, k3 + 1]:
            for m in range(18, 25):
                cells_to_border_1.append(chr(64 + m) + str(l))

        for l in range(5, k3 + 1):
            for m in ["R", "S", "Y"]:
                cells_to_border_2.append(m + str(l))

        for l in range(8, k3 + 1):
            cells_to_border_2.append("V" + str(l))

        for m in range(18, 25):
            cells_to_border_3.append(chr(64 + m) + str(i2))
            cells_to_border_3.append(chr(64 + m) + str(j2))

        # Adding border lines
        for cell in cells_to_border_1:
            sheet[cell].border = border_1

        for cell in cells_to_border_2:
            existing_border = sheet[cell].border
            left_border = existing_border.left
            right_border = existing_border.right
            top_border = existing_border.top
            bottom_border = existing_border.bottom
            new_border = Border(top=top_border, bottom=bottom_border,
                                left=border_2.left, right=right_border)
            sheet[cell].border = new_border

        for cell in cells_to_border_3:
            existing_border = sheet[cell].border
            left_border = existing_border.left
            right_border = existing_border.right
            top_border = existing_border.top
            bottom_border = existing_border.bottom
            new_border = Border(top=top_border, bottom=border_3.bottom,
                                left=left_border, right=right_border)
            sheet[cell].border = new_border

        for column in sheet.columns:  # Adjust cell dimension
            sheet.column_dimensions[column[0].column_letter].width = 12
            sheet.column_dimensions["B"].width = 40
            sheet.column_dimensions["J"].width = 40
            sheet.column_dimensions["R"].width = 40

        workbook.save(folder_path + file_name_xlsx)


def select_files():
    files, _ = QFileDialog.getOpenFileNames(None, "QFileDialog.getOpenFileNames()", "", "CSV files (*.csv)")
    if files:
        selectedFiles.append(";".join(files))
        for i in range(len(files)):
            Files.append(files[i])


def cancel():  # Cancel button
    window.close()


def confirm():  # Confirm button
    window.close()
    process_files()


app = QApplication(sys.argv)

window = QMainWindow()
window.setWindowTitle("File selectors")
window.setGeometry(400, 400, 500, 300)

layout = QVBoxLayout()

selectedFilesLabel = QLabel()
selectedFilesLabel.setText("Selected Files:")
layout.addWidget(selectedFilesLabel)

selectedFiles = QTextEdit()
selectedFiles.setReadOnly(True)
layout.addWidget(selectedFiles)

buttonBrowse = QPushButton("Browse")
buttonBrowse.clicked.connect(select_files)
layout.addWidget(buttonBrowse)

buttonLayout = QHBoxLayout()

buttonConfirm = QPushButton("Confirm")
buttonConfirm.clicked.connect(confirm)
buttonLayout.addWidget(buttonConfirm)

buttonCancel = QPushButton("Cancel")
buttonCancel.clicked.connect(cancel)
buttonLayout.addWidget(buttonCancel)

layout.addLayout(buttonLayout)

widget = QWidget()
widget.setLayout(layout)
window.setCentralWidget(widget)

window.show()
sys.exit(app.exec_())
