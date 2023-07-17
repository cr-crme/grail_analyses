# -*- coding: utf-8 -*-
"""
Created on Thu Jul 6 14:57:52 2023

@author: Florence
"""

import sys
import math
import os
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QTextEdit, QVBoxLayout, QHBoxLayout, QWidget, QLabel
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment, Border, Side 
global Files
Files = []

def columnAverage(fileData, variable, side, length):
    data = fileData[fileData.iloc[:, 1] == variable]
    ExcelData = {}

    if side == "left":
        dataLeft = data[data.iloc[:, 0] == "left"]
        columnAverageLeft = dataLeft.iloc[:, 3:].mean()

        leftOscillation = columnAverageLeft[length[0]:]
        maximumLeftOscillation = max(leftOscillation)
        minimumLeftOscillation = min(leftOscillation)
        rangeLeftOscillation = maximumLeftOscillation - minimumLeftOscillation

        leftAppui = columnAverageLeft[:length[0]]
        maximumLeftAppui = max(leftAppui)
        minimumLeftAppui = min(leftAppui)
        rangeLeftAppui = maximumLeftAppui - minimumLeftAppui

        ExcelData = {
            "Maximum Oscillation": maximumLeftOscillation,
            "Minimum Oscillation": minimumLeftOscillation,
            "Range Oscillation": rangeLeftOscillation,
            "Maximum Appui": maximumLeftAppui,
            "Minimum Appui": minimumLeftAppui,
            "Range Appui": rangeLeftAppui,
        }

    if side == "right":
        dataRight = data[data.iloc[:, 0] == "right"]
        columnAverageRight = dataRight.iloc[:, 3:].mean()

        rightOscillation = columnAverageRight[length[1]:]
        maximumRightOscillation = max(rightOscillation)
        minimumRightOscillation = min(rightOscillation)
        rangeRightOscillation = maximumRightOscillation - minimumRightOscillation

        rightAppui = columnAverageRight[:length[1]]
        maximumRightAppui = max(rightAppui)
        minimumRightAppui = min(rightAppui)
        rangeRightAppui = maximumRightAppui - minimumRightAppui

        ExcelData = {
            "Maximum Oscillation": maximumRightOscillation,
            "Minimum Oscillation": minimumRightOscillation,
            "Range Oscillation": rangeRightOscillation,
            "Maximum Appui": maximumRightAppui,
            "Minimum Appui": minimumRightAppui,
            "Range Appui": rangeRightAppui
        }
    return ExcelData

def process_files():
    for file in Files:
        ExcelData = {}
        fileData = pd.read_csv(file, header=None)
        
        # Mouvements sagittaux
        cinematicSagittal = ["Tronc gauche",
                              "Tronc droit",
                              "Bassin gauche",
                              "Bassin droit",
                              "Hanche gauche",
                              "Hanche droite",
                              "Genou gauche",
                              "Genou droit",
                              "Cheville gauche",
                              "Cheville droite",
                              ]
    
        cinematicSagittalEnglish = ["Rotation TrunkFlex",
                                    "Rotation TrunkFlex",
                                    "Rotation PelvicTil",
                                    "Rotation PelvicTil",
                                    "Rotation LHipFlex",
                                    "Rotation RHipFlex",
                                    "Rotation LKneeFlex",
                                    "Rotation RKneeFlex",
                                    "Rotation LAnkleFlex",
                                    "Rotation RAnkleFlex",
                                    ]
    
        momentSagittal = ["Hanche gauche",
                          "Hanche droite",
                          "Genou gauche",
                          "Genou droit",
                          "Cheville gauche",
                          "Cheville droite",
                          ]
        
        momentSagittalEnglish = ["Moment LHipFlex",
                                  "Moment RHipFlex",
                                  "Moment LKneeFlex",
                                  "Moment RKneeFlex",
                                  "Moment LAnkleFlex",
                                  "Moment RAnkleFlex",
                              ]
    
        powerSagittal = ["Hanche gauche",
                          "Hanche droite",
                          "Genou gauche",
                          "Genou droit",
                          "Cheville gauche",
                          "Cheville droite",
                          ]
        
        powerSagittalEnglish = ["Power LHipFlex",
                                "Power RHipFlex",
                                "Power LKneeFlex",
                                "Power RKneeFlex",
                                "Power LAnkleFlex",
                                "Power RAnkleFlex",
                                ]
    
        # Mouvements frontaux
        cinematicFrontal = ["Tronc gauche",
                            "Tronc droit",
                            "Bassin gauche",
                            "Bassin droit",
                            "Hanche gauche",
                            "Hanche droite",
                            "Genou gauche",
                            "Genou droit",
                            ]
        
        cinematicFrontalEnglish = ["Rotation TrunkTilt",
                                    "Rotation TrunkTilt",
                                    "Rotation PelvicObl",
                                    "Rotation PelvicObl",
                                    "Rotation LHipAbAd",
                                    "Rotation RHipAbAd",
                                    "Rotation LKneeAbAd",
                                    "Rotation RKneeAbAd",
                                    ]
        
        momentFrontal = ["Hanche gauche",
                          "Hanche droite",
                          "Genou gauche",
                          "Genou droit",
                          "Cheville gauche",
                          "Cheville droite",
                          ]
        
        momentFrontalEnglish = ["Moment LHipAbAd",
                                "Moment RHipAbAd",
                                "Moment LKneeAbAd",
                                "Moment RKneeAbAd",
                                "Moment LAnkleAbAd",
                                "Moment RAnkleAbAd",
                                ]
        
        powerFrontal = ["Hanche gauche",
                          "Hanche droite",
                          "Genou gauche",
                          "Genou droit",
                          "Cheville gauche",
                          "Cheville droite",
                          ]
    
        powerFrontalEnglish = ["Power LHipAbAd",
                                "Power RHipAbAd",
                                "Power LKneeAbAd",
                                "Power RKneeAbAd",
                                "Power LAnkleAbAd",
                                "Power RAnkleAbAd",
                                ]
    
        # Mouvements transversaux
        cinematicTransversal = ["Tronc gauche",
                                "Tronc droit",
                                "Bassin gauche",
                                "Bassin droit",
                                "Hanche gauche",
                                "Hanche droite",
                                "Genou gauche",
                                "Genou droit",
                                "Progression du pied gauche",
                                "Progression du pied droit",
                                ]
        
        cinematicTransversalEnglish = ["Rotation TrunkRot",
                                        "Rotation TrunkRot",
                                        "Rotation PelvicRot",
                                        "Rotation PelvicRot",
                                        "Rotation LHipRot",
                                        "Rotation RHipRot",
                                        "Rotation LKneeRot",
                                        "Rotation RKneeRot",
                                        "LProgression",
                                        "RProgression",
                                        ]
    
        momentTransversal = ["Hanche gauche",
                          "Hanche droite",
                          "Genou gauche",
                          "Genou droit",
                          "Cheville gauche",
                          "Cheville droite",
                          ]
        
        momentTransversalEnglish = ["Moment LHipRot",
                                    "Moment RHipRot",
                                    "Moment LKneeRot",
                                    "Moment RKneeRot",
                                    "Moment LAnkleRot",
                                    "Moment RAnkleRot",
                                    ]
        
        powerTransversal = ["Hanche gauche",
                          "Hanche droite",
                          "Genou gauche",
                          "Genou droit",
                          "Cheville gauche",
                          "Cheville droite",
                          ]
        
        powerTransversalEnglish = ["Power LHipRot",
                                    "Power RHipRot",
                                    "Power LKneeRot",
                                    "Power RKneeRot",
                                    "Power LAnkleRot",
                                    "Power RAnkleRot",
                                    ]
    
        
        # Déterminer nombre colonnes appartenant au swing et au stance
        # côté gauche
        LSwingTime = fileData[fileData.iloc[:, 1] == "L.Swing.Time"]
        LStanceTime = fileData[fileData.iloc[:, 1] == "L.Stance.Time"]
    
        leftLSwingTime = LSwingTime[LSwingTime.iloc[:, 0] == "left"]
        leftLStanceTime = LStanceTime[LStanceTime.iloc[:, 0] == "left"]
    
        leftLSwingTimeAverage = leftLSwingTime.iloc[:, 3].mean()
        leftLStanceTimeAverage = leftLStanceTime.iloc[:, 3].mean()
    
        totalLeftTimeAverage = leftLSwingTimeAverage + leftLStanceTimeAverage
        LSwingPercent = leftLSwingTimeAverage / totalLeftTimeAverage * 100
    
        # côté droit
        RSwingTime = fileData[fileData.iloc[:, 1] == "R.Swing.Time"]
        RStanceTime = fileData[fileData.iloc[:, 1] == "R.Stance.Time"]
    
        rightRSwingTime = RSwingTime[RSwingTime.iloc[:, 0] == "right"]
        rightRStanceTime = RStanceTime[RStanceTime.iloc[:, 0] == "right"]
    
        rightRSwingTimeAverage = rightRSwingTime.iloc[:, 3].mean()
        rightRStanceTimeAverage = rightRStanceTime.iloc[:, 3].mean()
    
        totalRightTimeAverage = rightRSwingTimeAverage + rightRStanceTimeAverage
        RSwingPercent = rightRSwingTimeAverage / totalRightTimeAverage * 100
    
        # Produit croisé pour déterminer le nombre de colonnes faisant partie du Swing (sur 101 colonnes totales)
        nbLeftSwing = math.ceil(LSwingPercent * 101 / 100)
        nbRightSwing = math.ceil(RSwingPercent * 101 / 100)
        
        
        # cinematic
        for i in range(0, len(cinematicSagittalEnglish), 2):
            ExcelData[cinematicSagittalEnglish[i]] = columnAverage(fileData, cinematicSagittalEnglish[i], "left", [
                          nbLeftSwing, nbRightSwing])

        for i in range(0, len(cinematicFrontalEnglish), 2):
            ExcelData[cinematicFrontalEnglish[i]] = columnAverage(fileData, cinematicFrontalEnglish[i], "left", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(0, len(cinematicTransversalEnglish), 2):
            ExcelData[cinematicTransversalEnglish[i]] = columnAverage(fileData, cinematicTransversalEnglish[i], "left", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(cinematicSagittalEnglish), 2):
            ExcelData[cinematicSagittalEnglish[i]] = columnAverage(fileData, cinematicSagittalEnglish[i], "right", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(cinematicFrontalEnglish), 2):
            ExcelData[cinematicFrontalEnglish[i]] = columnAverage(fileData, cinematicFrontalEnglish[i], "right", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(cinematicTransversalEnglish), 2):
            ExcelData[cinematicTransversalEnglish[i]] = columnAverage(fileData, cinematicTransversalEnglish[i], "right", [
                          nbLeftSwing, nbRightSwing])
    
        # moment
        for i in range(0, len(momentSagittalEnglish), 2):
            ExcelData[momentSagittalEnglish[i]] = columnAverage(fileData, momentSagittalEnglish[i], "left", [nbLeftSwing, nbRightSwing])
    
        for i in range(0, len(momentFrontalEnglish), 2):
            ExcelData[momentFrontalEnglish[i]] = columnAverage(fileData, momentFrontalEnglish[i], "left", [nbLeftSwing, nbRightSwing])
    
        for i in range(0, len(momentTransversalEnglish), 2):
            ExcelData[momentTransversalEnglish[i]] = columnAverage(fileData, momentTransversalEnglish[i], "left", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(momentSagittalEnglish), 2):
            ExcelData[momentSagittalEnglish[i]] = columnAverage(fileData, momentSagittalEnglish[i], "right", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(momentFrontalEnglish), 2):
            ExcelData[momentFrontalEnglish[i]] = columnAverage(fileData, momentFrontalEnglish[i], "right", [nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(momentTransversalEnglish), 2):
            ExcelData[momentTransversalEnglish[i]] = columnAverage(fileData, momentTransversalEnglish[i], "right", [
                          nbLeftSwing, nbRightSwing])
    
        # power
        for i in range(0, len(powerSagittalEnglish), 2):
            ExcelData[powerSagittalEnglish[i]] = columnAverage(fileData, powerSagittalEnglish[i], "left", [nbLeftSwing, nbRightSwing])
    
        for i in range(0, len(powerFrontalEnglish), 2):
            ExcelData[powerFrontalEnglish[i]] = columnAverage(fileData, powerFrontalEnglish[i], "left", [nbLeftSwing, nbRightSwing])
    
        for i in range(0, len(powerTransversalEnglish), 2):
            ExcelData[powerTransversalEnglish[i]] = columnAverage(fileData, powerTransversalEnglish[i], "left", [
                          nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(powerSagittalEnglish), 2):
            ExcelData[powerSagittalEnglish[i]] = columnAverage(fileData, powerSagittalEnglish[i], "right", [nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(powerFrontalEnglish), 2):
            ExcelData[powerFrontalEnglish[i]] = columnAverage(fileData, powerFrontalEnglish[i], "right", [nbLeftSwing, nbRightSwing])
    
        for i in range(1, len(powerTransversalEnglish), 2):
            ExcelData[powerTransversalEnglish[i]] = columnAverage(fileData, powerTransversalEnglish[i], "right", [
                          nbLeftSwing, nbRightSwing])
    
        # PDF
        folderPath = "C:/Users/Florence/Desktop/STAGE/test/"
        fileNamePDF = os.path.splitext(os.path.basename(file))[0] + ".pdf"
        doc = SimpleDocTemplate(
            folderPath + fileNamePDF, pagesize=letter, leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        elements = []
        
        identification = [["ID", ""],
                          ["Session ID", ""],
                          ["Date", ""]
                          ]
        identificationTable = Table(identification,  colWidths=[75, 400])
        identificationTable.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
        elements.append(identificationTable)
        elements.append(Spacer(1, 20))  # Add an empty spacer for separation
        
        # Cinématique
        title = [["Cinématique"],
                  ]
        titleTable = Table(title, colWidths=[475])
        titleTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
        elements.append(titleTable)
        
        phases = [["", "Phase d'Appui", "Phase d'Oscillations"],
                  ]
        phasesTable = Table(phases, colWidths=[140, 183, 183])
        phasesTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
            ]))
        elements.append(phasesTable)
        
        measures = [["", "Min (deg)", "Max (deg)", "Range (deg)", "Min (deg)", "Max (deg)", "Range (deg)"],
                    ["Sagittal"],
                    ]
        for i in range(len(cinematicSagittal)):
            row = [cinematicSagittal[i], 
                     round(ExcelData[cinematicSagittalEnglish[i]]["Minimum Appui"], 2),
                    round(ExcelData[cinematicSagittalEnglish[i]]["Maximum Appui"], 2),
                    round(ExcelData[cinematicSagittalEnglish[i]]["Range Appui"], 2),
                    round(ExcelData[cinematicSagittalEnglish[i]]["Minimum Oscillation"], 2), 
                    round(ExcelData[cinematicSagittalEnglish[i]]["Maximum Oscillation"], 2),
                    round(ExcelData[cinematicSagittalEnglish[i]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measures.append(["Frontal"])
        for j in range(len(cinematicFrontal)):
            row = [cinematicFrontal[j], 
                    round(ExcelData[cinematicFrontalEnglish[j]]["Minimum Appui"], 2),
                    round(ExcelData[cinematicFrontalEnglish[j]]["Maximum Appui"], 2),
                    round(ExcelData[cinematicFrontalEnglish[j]]["Range Appui"], 2),
                    round(ExcelData[cinematicFrontalEnglish[j]]["Minimum Oscillation"], 2), 
                    round(ExcelData[cinematicFrontalEnglish[j]]["Maximum Oscillation"], 2),
                    round(ExcelData[cinematicFrontalEnglish[j]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measures.append(["Transversal"])
        for k in range(len(cinematicTransversal)):
            row = [cinematicTransversal[k], 
                    round(ExcelData[cinematicTransversalEnglish[k]]["Minimum Appui"], 2),
                    round(ExcelData[cinematicTransversalEnglish[k]]["Maximum Appui"], 2),
                    round(ExcelData[cinematicTransversalEnglish[k]]["Range Appui"], 2),
                    round(ExcelData[cinematicTransversalEnglish[k]]["Minimum Oscillation"], 2), 
                    round(ExcelData[cinematicTransversalEnglish[k]]["Maximum Oscillation"], 2),
                    round(ExcelData[cinematicTransversalEnglish[k]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measureTable = Table(measures, colWidths=[140, 61, 61, 61, 61, 61, 61])
        measureTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, i+3), (0, i+3), 'Helvetica-Bold'),
            ('FONTNAME', (0, i+j+5), (0, i+j+5), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (-4,0), (-4, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, i+2), (-1, i+2), 1, colors.black),
            ('LINEBELOW', (0, i+j+4), (-1, i+j+4), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
            ]))
        
        elements.append(measureTable)
        elements.append(PageBreak())
        
        # Moment
        title = [["Moment"],
                  ]
        titleTable = Table(title, colWidths=[475])
        titleTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
        elements.append(titleTable)
        
        phases = [["", "Phase d'Appui", "Phase d'Oscillations"],
                  ]
        phasesTable = Table(phases, colWidths=[140, 183, 183])
        phasesTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
            ]))
        elements.append(phasesTable)
        
        measures = [["", "Min (deg)", "Max (deg)", "Range (deg)", "Min (deg)", "Max (deg)", "Range (deg)"],
                    ["Sagittal"],
                    ]
        for i in range(len(momentSagittal)):
            row = [momentSagittal[i], 
                     round(ExcelData[momentSagittalEnglish[i]]["Minimum Appui"], 2),
                    round(ExcelData[momentSagittalEnglish[i]]["Maximum Appui"], 2),
                    round(ExcelData[momentSagittalEnglish[i]]["Range Appui"], 2),
                    round(ExcelData[momentSagittalEnglish[i]]["Minimum Oscillation"], 2), 
                    round(ExcelData[momentSagittalEnglish[i]]["Maximum Oscillation"], 2),
                    round(ExcelData[momentSagittalEnglish[i]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measures.append(["Frontal"])
        for j in range(len(momentFrontal)):
            row = [momentFrontal[j], 
                    round(ExcelData[momentFrontalEnglish[j]]["Minimum Appui"], 2),
                    round(ExcelData[momentFrontalEnglish[j]]["Maximum Appui"], 2),
                    round(ExcelData[momentFrontalEnglish[j]]["Range Appui"], 2),
                    round(ExcelData[momentFrontalEnglish[j]]["Minimum Oscillation"], 2), 
                    round(ExcelData[momentFrontalEnglish[j]]["Maximum Oscillation"], 2),
                    round(ExcelData[momentFrontalEnglish[j]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measures.append(["Transversal"])
        for k in range(len(momentTransversal)):
            row = [momentTransversal[k], 
                    round(ExcelData[momentTransversalEnglish[k]]["Minimum Appui"], 2),
                    round(ExcelData[momentTransversalEnglish[k]]["Maximum Appui"], 2),
                    round(ExcelData[momentTransversalEnglish[k]]["Range Appui"], 2),
                    round(ExcelData[momentTransversalEnglish[k]]["Minimum Oscillation"], 2), 
                    round(ExcelData[momentTransversalEnglish[k]]["Maximum Oscillation"], 2),
                    round(ExcelData[momentTransversalEnglish[k]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measureTable = Table(measures, colWidths=[140, 61, 61, 61, 61, 61, 61])
        measureTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, i+3), (0, i+3), 'Helvetica-Bold'),
            ('FONTNAME', (0, i+j+5), (0, i+j+5), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (-4,0), (-4, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, i+2), (-1, i+2), 1, colors.black),
            ('LINEBELOW', (0, i+j+4), (-1, i+j+4), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
            ]))
        
        elements.append(measureTable)
        elements.append(PageBreak())
        
        # Power
        title = [["Puissance"],
                  ]
        titleTable = Table(title, colWidths=[475])
        titleTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
        elements.append(titleTable)
        
        phases = [["", "Phase d'Appui", "Phase d'Oscillations"],
                  ]
        phasesTable = Table(phases, colWidths=[140, 183, 183])
        phasesTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (0, 0), (-1, -1), 1, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
            ]))
        elements.append(phasesTable)
        
        measures = [["", "Min (deg)", "Max (deg)", "Range (deg)", "Min (deg)", "Max (deg)", "Range (deg)"],
                    ["Sagittal"],
                    ]
        for i in range(len(powerSagittal)):
            row = [powerSagittal[i], 
                     round(ExcelData[powerSagittalEnglish[i]]["Minimum Appui"], 2),
                    round(ExcelData[powerSagittalEnglish[i]]["Maximum Appui"], 2),
                    round(ExcelData[powerSagittalEnglish[i]]["Range Appui"], 2),
                    round(ExcelData[powerSagittalEnglish[i]]["Minimum Oscillation"], 2), 
                    round(ExcelData[powerSagittalEnglish[i]]["Maximum Oscillation"], 2),
                    round(ExcelData[powerSagittalEnglish[i]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measures.append(["Frontal"])
        for j in range(len(powerFrontal)):
            row = [powerFrontal[j], 
                    round(ExcelData[powerFrontalEnglish[j]]["Minimum Appui"], 2),
                    round(ExcelData[powerFrontalEnglish[j]]["Maximum Appui"], 2),
                    round(ExcelData[powerFrontalEnglish[j]]["Range Appui"], 2),
                    round(ExcelData[powerFrontalEnglish[j]]["Minimum Oscillation"], 2), 
                    round(ExcelData[powerFrontalEnglish[j]]["Maximum Oscillation"], 2),
                    round(ExcelData[powerFrontalEnglish[j]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measures.append(["Transversal"])
        for k in range(len(powerTransversal)):
            row = [powerTransversal[k], 
                    round(ExcelData[powerTransversalEnglish[k]]["Minimum Appui"], 2),
                    round(ExcelData[powerTransversalEnglish[k]]["Maximum Appui"], 2),
                    round(ExcelData[powerTransversalEnglish[k]]["Range Appui"], 2),
                    round(ExcelData[powerTransversalEnglish[k]]["Minimum Oscillation"], 2), 
                    round(ExcelData[powerTransversalEnglish[k]]["Maximum Oscillation"], 2),
                    round(ExcelData[powerTransversalEnglish[k]]["Range Oscillation"], 2),
                    ]
            measures.append(row)
            
        measureTable = Table(measures, colWidths=[140, 61, 61, 61, 61, 61, 61])
        measureTable.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('FONTNAME', (0, i+3), (0, i+3), 'Helvetica-Bold'),
            ('FONTNAME', (0, i+j+5), (0, i+j+5), 'Helvetica-Bold'),
            ('LINEBEFORE', (0, 0), (0, -1), 2, colors.black),
            ('LINEAFTER', (-1, 0), (-1, -1), 2, colors.black),
            ('LINEAFTER', (-4,0), (-4, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 1, colors.black),
            ('LINEBELOW', (0, i+2), (-1, i+2), 1, colors.black),
            ('LINEBELOW', (0, i+j+4), (-1, i+j+4), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
            ]))
        
        elements.append(measureTable)
        elements.append(PageBreak())
        doc.build(elements)
        
        
        # EXCEL
        folderPath = "C:/Users/Florence/Desktop/STAGE/test/"
        fileNameXLSX = os.path.splitext(os.path.basename(file))[0] + ".xlsx"
        
        workbook = Workbook() 
        sheet = workbook.active 
        # Titre de cellules 
        sheet["B2"] = "Patient ID" 
        sheet["B3"] = "Session ID" 
        sheet["B4"] = "Date" 
        sheet["C5"] = "Cinématique" 
        sheet["C6"] = "Phase d'appui" 
        sheet["F6"] = "Phase d'oscillation" 
        sheet["C7"] = "Min (deg)" 
        sheet["D7"] = "Max (deg)" 
        sheet["E7"] = "Range(deg)" 
        sheet["F7"] = "Min (deg)" 
        sheet["G7"] = "Max (deg)" 
        sheet["H7"] = "Range(deg)" 
         
        sheet["K5"] = "Moment" 
        sheet["K6"] = "Phase d'appui" 
        sheet["N6"] = "Phase d'oscillation" 
        sheet["K7"] = "Min (deg)" 
        sheet["L7"] = "Max (deg)" 
        sheet["M7"] = "Range(deg)" 
        sheet["N7"] = "Min (deg)" 
        sheet["O7"] = "Max (deg)" 
        sheet["P7"] = "Range(deg)" 
         
        sheet["S5"] = "Power" 
        sheet["S6"] = "Phase d'appui" 
        sheet["V6"] = "Phase d'oscillation" 
        sheet["S7"] = "Min (deg)" 
        sheet["T7"] = "Max (deg)" 
        sheet["U7"] = "Range(deg)" 
        sheet["V7"] = "Min (deg)" 
        sheet["W7"] = "Max (deg)" 
        sheet["X7"] = "Range(deg)" 
         
        # Titre des variables 
         
        # Cinématique 
        i1 = 8 
        sheet["B"+str(i1)] = "Sagittal" 
         
        count = 0 
        for i1, value in enumerate(cinematicSagittal, start=i1+1): 
            sheet["B" + str(i1)] = value 
            sheet["C" + str(i1)] = round(ExcelData[cinematicSagittalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["D" + str(i1)] = round(ExcelData[cinematicSagittalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["E" + str(i1)] = round(ExcelData[cinematicSagittalEnglish[count]] 
                                         ["Range Appui"], 2) 
            sheet["F" + str(i1)] = round(ExcelData[cinematicSagittalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["G" + str(i1)] = round(ExcelData[cinematicSagittalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["H" + str(i1)] = round(ExcelData[cinematicSagittalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        j1 = i1 + 1 
        sheet["B"+str(j1)] = "Frontal" 
         
        count = 0 
        for j1, value in enumerate(cinematicFrontal, start=(j1+1)): 
            sheet["B" + str(j1)] = value 
            sheet["C" + str(j1)] = round(ExcelData[cinematicFrontalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["D" + str(j1)] = round(ExcelData[cinematicFrontalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["E" + str(j1)] = round(ExcelData[cinematicFrontalEnglish[count]] 
                                         ["Range Appui"], 2) 
            sheet["F" + str(j1)] = round(ExcelData[cinematicFrontalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["G" + str(j1)] = round(ExcelData[cinematicFrontalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["H" + str(j1)] = round(ExcelData[cinematicFrontalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        k1 = j1 + 1 
        sheet["B"+str(k1)] = "Transversal" 
         
        count = 0 
        for k1, value in enumerate(cinematicTransversal, start=(k1+1)): 
            sheet["B" + str(k1)] = value 
            sheet["C" + str(k1)] = round(ExcelData[cinematicTransversalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["D" + str(k1)] = round(ExcelData[cinematicTransversalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["E" + str(k1)] = round(ExcelData[cinematicTransversalEnglish[count]] 
                                         ["Range Appui"], 2) 
            sheet["F" + str(k1)] = round(ExcelData[cinematicTransversalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["G" + str(k1)] = round(ExcelData[cinematicTransversalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["H" + str(k1)] = round(ExcelData[cinematicTransversalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        # Moment 
        i2 = 8 
        sheet["J"+str(i2)] = "Sagittal" 
         
        count = 0 
        for i2, value in enumerate(momentSagittal, start=i2+1): 
            sheet["J" + str(i2)] = value 
            sheet["K" + str(i2)] = round(ExcelData[momentSagittalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["L" + str(i2)] = round(ExcelData[momentSagittalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["M" + str(i2)] = round(ExcelData[momentSagittalEnglish[count]] 
                                         ["Range Appui"], 2) 
            sheet["N" + str(i2)] = round(ExcelData[momentSagittalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["O" + str(i2)] = round(ExcelData[momentSagittalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["P" + str(i2)] = round(ExcelData[momentSagittalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        j2 = i2 + 1 
        sheet["J"+str(j2)] = "Frontal" 
         
        count = 0 
        j2 = i2 + 1 
        for j2, value in enumerate(momentFrontal, start=(j2+1)): 
            sheet["J" + str(j2)] = value 
            sheet["K" + str(j2)] = round(ExcelData[momentFrontalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["L" + str(j2)] = round(ExcelData[momentFrontalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["M" + str(j2)] = round(ExcelData[momentFrontalEnglish[count]]["Range Appui"], 2) 
            sheet["N" + str(j2)] = round(ExcelData[momentFrontalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["O" + str(j2)] = round(ExcelData[momentFrontalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["P" + str(j2)] = round(ExcelData[momentFrontalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        k2 = j2 + 1 
        sheet["J"+str(k2)] = "Transversal" 
         
        count = 0 
        for k2, value in enumerate(momentTransversal, start=(k2+1)): 
            sheet["J" + str(k2)] = value 
            sheet["K" + str(k2)] = round(ExcelData[momentTransversalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["L" + str(k2)] = round(ExcelData[momentTransversalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["M" + str(k2)] = round(ExcelData[momentTransversalEnglish[count]] 
                                         ["Range Appui"], 2) 
            sheet["N" + str(k2)] = round(ExcelData[momentTransversalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["O" + str(k2)] = round(ExcelData[momentTransversalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["P" + str(k2)] = round(ExcelData[momentTransversalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        # Power 
        i3 = 8 
        sheet["R"+str(i3)] = "Sagittal" 
         
        count = 0 
        for i3, value in enumerate(powerSagittal, start=i3+1): 
            sheet["R" + str(i3)] = value 
            sheet["S" + str(i3)] = round(ExcelData[powerSagittalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["T" + str(i3)] = round(ExcelData[powerSagittalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["U" + str(i3)] = round(ExcelData[powerSagittalEnglish[count]]["Range Appui"], 2) 
            sheet["V" + str(i3)] = round(ExcelData[powerSagittalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["W" + str(i3)] = round(ExcelData[powerSagittalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["X" + str(i3)] = round(ExcelData[powerSagittalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        j3 = i3 + 1 
        sheet["R"+str(j3)] = "Frontal" 
         
        count = 0 
        j3 = i3 + 1 
        for j3, value in enumerate(powerFrontal, start=(j3+1)): 
            sheet["R" + str(j3)] = value 
            sheet["S" + str(j3)] = round(ExcelData[powerFrontalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["T" + str(j3)] = round(ExcelData[powerFrontalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["U" + str(j3)] = round(ExcelData[powerFrontalEnglish[count]]["Range Appui"], 2) 
            sheet["V" + str(j3)] = round(ExcelData[powerFrontalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["W" + str(j3)] = round(ExcelData[powerFrontalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["X" + str(j3)] = round(ExcelData[powerFrontalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        k3 = j3 + 1 
        sheet["R"+str(k3)] = "Transversal" 
         
        count = 0 
        for k3, value in enumerate(powerTransversal, start=(k3 + 1)): 
            sheet["R" + str(k3)] = value 
            sheet["S" + str(k3)] = round(ExcelData[powerTransversalEnglish[count]] 
                                         ["Minimum Appui"], 2) 
            sheet["T" + str(k3)] = round(ExcelData[powerTransversalEnglish[count]] 
                                         ["Maximum Appui"], 2) 
            sheet["U" + str(k3)] = round(ExcelData[powerTransversalEnglish[count]] 
                                         ["Range Appui"], 2) 
            sheet["V" + str(k3)] = round(ExcelData[powerTransversalEnglish[count]] 
                                         ["Minimum Oscillation"], 2) 
            sheet["W" + str(k3)] = round(ExcelData[powerTransversalEnglish[count]] 
                                         ["Maximum Oscillation"], 2) 
            sheet["X" + str(k3)] = round(ExcelData[powerTransversalEnglish[count]] 
                                         ["Range Oscillation"], 2) 
            count += 1 
         
        # Mise en forme 
        sheet.merge_cells("C2:H2")  # Merge de cellules 
        sheet.merge_cells("C3:H3") 
        sheet.merge_cells("C4:H4") 
        sheet.merge_cells("C5:H5") 
        sheet.merge_cells("K5:P5") 
        sheet.merge_cells("S5:X5") 
        sheet.merge_cells("C6:E6") 
        sheet.merge_cells("F6:H6") 
        sheet.merge_cells("K6:M6") 
        sheet.merge_cells("N6:P6") 
        sheet.merge_cells("S6:U6") 
        sheet.merge_cells("V6:X6") 
         
        sheet["C5"].font = Font(bold=True)  # Mise de titres en gras 
        sheet["K5"].font = Font(bold=True) 
        sheet["S5"].font = Font(bold=True) 
        sheet["C6"].font = Font(bold=True) 
        sheet["F6"].font = Font(bold=True) 
        sheet["K6"].font = Font(bold=True) 
        sheet["N6"].font = Font(bold=True) 
        sheet["S6"].font = Font(bold=True) 
        sheet["V6"].font = Font(bold=True) 
        sheet["B8"].font = Font(bold=True)           # Sagittal 
        sheet["J8"].font = Font(bold=True) 
        sheet["R8"].font = Font(bold=True) 
        sheet["B"+str(i1+1)].font = Font(bold=True)  # Frontal 
        sheet["J"+str(i2+1)].font = Font(bold=True) 
        sheet["R"+str(i3+1)].font = Font(bold=True) 
        sheet["B"+str(j1+1)].font = Font(bold=True)  # Transversal 
        sheet["J"+str(j2+1)].font = Font(bold=True) 
        sheet["R"+str(j3+1)].font = Font(bold=True) 
         
         
        # Centrer le texte dansle cellules 
        alignment = Alignment(horizontal="center", vertical="center") 
        for row in sheet.iter_rows(): 
            for cell in row: 
                cell.alignment = alignment 
         
         
        cells_to_border_1 = []  # Bordures 
        cells_to_border_2 = [] 
        cells_to_border_3 = [] 
        border_1 = Border(top=Side(border_style="thick", color="000000")) 
        border_2 = Border(left=Side(border_style="thick", color="000000")) 
        border_3 = Border(bottom=Side(border_style="medium", color="000000")) 
         
        # Premier tableau 
        for l in [2, 5, 8, k1+1]: 
            for m in range(2, 9): 
                cells_to_border_1.append(chr(64+m) + str(l)) 
         
        for l in range(2, k1+1): 
            for m in ["B", "I"]: 
                cells_to_border_2.append(m + str(l)) 
         
        for l in range(5, k1+1): 
            cells_to_border_2.append("C" + str(l)) 
         
        for l in range(8, k1+1): 
            cells_to_border_2.append("F" + str(l)) 
         
        for m in range(2, 9): 
            cells_to_border_3.append(chr(64+m) + str(i1)) 
            cells_to_border_3.append(chr(64+m) + str(j1)) 
         
        # Deuxième tableau 
        for l in [5, 8, k2+1]: 
            for m in range(10, 17): 
                cells_to_border_1.append(chr(64+m) + str(l)) 
         
        for l in range(5, k2+1): 
            for m in ["J", "K", "Q"]: 
                cells_to_border_2.append(m + str(l)) 
         
        for l in range(8, k2+1): 
            cells_to_border_2.append("N" + str(l)) 
         
        for m in range(10, 17): 
            cells_to_border_3.append(chr(64+m) + str(i2)) 
            cells_to_border_3.append(chr(64+m) + str(j2)) 
         
        # Troisième tableau 
        for l in [5, 8, k3+1]: 
            for m in range(18, 25): 
                cells_to_border_1.append(chr(64+m) + str(l)) 
         
        for l in range(5, k3+1): 
            for m in ["R", "S", "Y"]: 
                cells_to_border_2.append(m + str(l)) 
         
        for l in range(8, k3+1): 
            cells_to_border_2.append("V" + str(l)) 
         
        for m in range(18, 25): 
            cells_to_border_3.append(chr(64+m) + str(i2)) 
            cells_to_border_3.append(chr(64+m) + str(j2)) 
         
        # Réalisation des lignes de contour 
        for cell in cells_to_border_1: 
            sheet[cell].border = border_1 
         
        for cell in cells_to_border_2: 
            existing_border = sheet[cell].border 
            left_border = existing_border.left 
            right_border = existing_border.right 
            top_border = existing_border.top 
            bottom_border = existing_border.bottom 
            new_border = Border(top=top_border, bottom=bottom_border, 
                                left=border_2.left, right=right_border) 
            sheet[cell].border = new_border 
         
        for cell in cells_to_border_3: 
            existing_border = sheet[cell].border 
            left_border = existing_border.left 
            right_border = existing_border.right 
            top_border = existing_border.top 
            bottom_border = existing_border.bottom 
            new_border = Border(top=top_border, bottom=border_3.bottom, 
                                left=left_border, right=right_border) 
            sheet[cell].border = new_border 
         
        for column in sheet.columns:  # Ajuster la taille des cellules 
            sheet.column_dimensions[column[0].column_letter].width = 12 
            sheet.column_dimensions["B"].width = 40 
            sheet.column_dimensions["J"].width = 40 
            sheet.column_dimensions["R"].width = 40 
         
        workbook.save(folderPath + fileNameXLSX) 



def select_files():
    files, _ = QFileDialog.getOpenFileNames(None, "QFileDialog.getOpenFileNames()", "", "CSV files (*.csv)")
    if files:
        selectedFiles.append(";".join(files))
        for i in range(len(files)):
            Files.append(files[i])

def clear_files():
    selectedFiles.clear()
    window.close()

def confirm_files():
    window.close()
    selectedFiles.clear()
    process_files()
    

app = QApplication(sys.argv)

window = QMainWindow()
window.setWindowTitle("File selectors")
window.setGeometry(400, 400, 500, 300)

layout = QVBoxLayout()

selectedFilesLabel = QLabel()
selectedFilesLabel.setText("Selected Files:")
layout.addWidget(selectedFilesLabel)

selectedFiles = QTextEdit()
selectedFiles.setReadOnly(True)
layout.addWidget(selectedFiles)

buttonBrowse = QPushButton("Browse")
buttonBrowse.clicked.connect(select_files)
layout.addWidget(buttonBrowse)

buttonLayout = QHBoxLayout()

buttonConfirm = QPushButton("Confirm")
buttonConfirm.clicked.connect(confirm_files)
buttonLayout.addWidget(buttonConfirm)

buttonCancel = QPushButton("Cancel")
buttonCancel.clicked.connect(clear_files)
buttonLayout.addWidget(buttonCancel)

layout.addLayout(buttonLayout)

widget = QWidget()
widget.setLayout(layout)
window.setCentralWidget(widget)

window.show()
sys.exit(app.exec_())